from openpyxl import Workbook # Para instalar: pip install openpyxl

# Criando uma planilha Excel
wb = Workbook()
ws = wb.active
ws.title = "Funcionários"

# Adicionando dados
ws.append(["Nome", "Cargo", "Salário"])
ws.append(["Maria", "Engenheira", 8500])
ws.append(["Pedro", "Analista", 6000])

# Salvando o arquivo
wb.save("funcionarios.xlsx")

# Lendo o arquivo
from openpyxl import load_workbook
wb = load_workbook("funcionarios.xlsx")
ws = wb.active
for linha in ws.iter_rows(values_only=True):
    print(linha)

#EXTRA: lendo xlsx com pandas
import pandas as pd # Para instalar: pip install pandas

df = pd.read_excel('funcionarios.xlsx')
print(df.head())  # Exibe as primeiras linhas